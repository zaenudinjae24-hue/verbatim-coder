import streamlit as st
import pandas as pd
import json
import os
import time
import re
import requests
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="AI Verbatim Coder",
    page_icon="🧠",
    layout="wide"
)

MEMORY_BANK_PATH = "memory_bank.json"
GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def load_memory_bank():
    if os.path.exists(MEMORY_BANK_PATH):
        with open(MEMORY_BANK_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"studies": [], "coding_rules": {}}


def save_memory_bank(mb):
    with open(MEMORY_BANK_PATH, "w", encoding="utf-8") as f:
        json.dump(mb, f, ensure_ascii=False, indent=2)


def call_gemini(api_key: str, prompt: str, temperature: float = 0.2) -> str:
    headers = {"Content-Type": "application/json"}
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": temperature, "maxOutputTokens": 8192}
    }
    resp = requests.post(
        f"{GEMINI_API_URL}?key={api_key}",
        headers=headers,
        json=payload,
        timeout=120
    )
    if resp.status_code != 200:
        raise Exception(f"Gemini API error {resp.status_code}: {resp.text[:300]}")
    data = resp.json()
    return data["candidates"][0]["content"]["parts"][0]["text"]


def build_memory_context(mb: dict) -> str:
    """Ringkas memory bank jadi string konteks untuk prompt."""
    rules = mb.get("coding_rules", {})
    lines = ["=== MEMORY BANK: Pola Coding dari Studi Sebelumnya ===\n"]

    if rules:
        lines.append("ATURAN CODING:")
        lines.append(f"- Separator multi-code: {rules.get('multi_code_separator', ';')}")
        conv = rules.get("numbering_convention", {})
        for k, v in conv.items():
            lines.append(f"- {k}: {v}")
        lines.append(f"- Format Nett: {rules.get('nett_format', '')}\n")

    for study in mb.get("studies", []):
        lines.append(f"--- STUDI: {study['study_id']} ---")
        lines.append(f"Topik: {study.get('topic', '')}")
        lines.append(f"Pertanyaan: {study.get('question', '')}")
        lines.append(f"Tipe: {study.get('type', '')}")

        codelist = study.get("codelist", [])
        if codelist:
            lines.append("Contoh kode yang digunakan:")
            for c in codelist[:20]:
                nett = f"[{c['nett']}] " if c.get('nett') else ""
                lines.append(f"  {c['code']} = {nett}{c['label_id']} / {c['label_en']}")
            if len(codelist) > 20:
                lines.append(f"  ... (total {len(codelist)} kode)")

        samples = study.get("verbatim_samples", [])
        if samples:
            lines.append("Contoh verbatim → kode:")
            for s in samples[:15]:
                lines.append(f"  \"{s['verbatim']}\" → {s['code']}")
        lines.append("")

    return "\n".join(lines)


def build_codeframe_prompt(mb: dict, verbatims: list, question_context: str, language: str) -> str:
    memory_ctx = build_memory_context(mb)
    verbatim_sample = "\n".join([f"- {v}" for v in verbatims[:80]])

    lang_instruction = {
        "Bilingual (ID + EN)": "Buat label dalam dua bahasa: Indonesia (label_id) dan Inggris (label_en).",
        "Indonesia saja": "Buat label dalam Bahasa Indonesia saja (label_en sama dengan label_id).",
        "Inggris saja": "Buat label dalam Bahasa Inggris saja (label_id sama dengan label_en)."
    }.get(language, "Buat label dalam dua bahasa.")

    return f"""Kamu adalah ahli coding verbatim riset pasar dengan pengalaman luas.

{memory_ctx}

=== TUGAS SEKARANG ===
Pertanyaan/Konteks: {question_context}
Bahasa: {lang_instruction}

Berikut adalah sample verbatim dari responden:
{verbatim_sample}

INSTRUKSI:
1. Baca semua verbatim di atas dengan cermat.
2. Identifikasi tema-tema utama yang muncul.
3. Buat CODEFRAME (daftar kode) yang mencakup semua tema penting.
4. Kelompokkan kode ke dalam NETT (kategori besar) yang sesuai.
5. Ikuti konvensi penomoran dari Memory Bank (misal: positif=100-an, negatif=500-an, brand=1-99).
6. Setiap kode harus spesifik namun tidak terlalu granular (jangan buat kode untuk hal yang hanya muncul 1x kecuali sangat unik).
7. Selalu sertakan kode 999 = Tidak ada / Sudah bagus (jika relevan) dan 99 = Lainnya.

Kembalikan HANYA JSON berikut (tanpa teks lain, tanpa markdown):
{{
  "codeframe": [
    {{
      "code": 101,
      "nett": "Nama Nett/Kategori",
      "label_id": "Label Indonesia",
      "label_en": "Label English"
    }}
  ],
  "suggested_netts": ["Nett 1", "Nett 2", "Nett 3"]
}}"""


def build_autocode_prompt(mb: dict, codeframe: list, verbatims_batch: list, question_context: str) -> str:
    code_list = "\n".join([
        f"  {c['code']} = [{c['nett']}] {c['label_id']} / {c['label_en']}"
        for c in codeframe
    ])

    verb_list = "\n".join([
        f"  {i+1}. \"{v}\""
        for i, v in enumerate(verbatims_batch)
    ])

    samples_ctx = ""
    for study in mb.get("studies", [])[:2]:
        s = study.get("verbatim_samples", [])[:8]
        if s:
            samples_ctx += f"\nContoh dari studi {study['study_id']}:\n"
            for ex in s:
                samples_ctx += f"  \"{ex['verbatim']}\" → {ex['code']}\n"

    return f"""Kamu adalah ahli coding verbatim riset pasar.

Pertanyaan: {question_context}

CODEFRAME YANG DIGUNAKAN:
{code_list}

ATURAN:
- Separator multi-code: ; (contoh: 101;165 jika verbatim menyebut dua tema)
- Null/kosong/tidak relevan: biarkan kosong
- Jika tidak ada yang cocok: gunakan 99
- Tidak ada saran / sudah bagus: gunakan 999
- Satu verbatim BISA dapat lebih dari 1 kode jika memang menyebut beberapa hal{samples_ctx}

VERBATIM YANG HARUS DI-CODE ({len(verbatims_batch)} baris):
{verb_list}

Kembalikan HANYA JSON array berikut (tanpa teks lain, tanpa markdown):
[
  {{"row": 1, "verbatim": "...", "code": "101"}},
  {{"row": 2, "verbatim": "...", "code": "101;165"}},
  {{"row": 3, "verbatim": "", "code": ""}},
  ...
]
Jumlah item dalam array harus tepat {len(verbatims_batch)}."""


def parse_json_response(text: str):
    text = text.strip()
    text = re.sub(r'^```json\s*', '', text)
    text = re.sub(r'^```\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    return json.loads(text)


def create_output_excel(df_result: pd.DataFrame, codeframe: list, col_groups: list) -> BytesIO:
    wb = Workbook()

    # Sheet 1: Codeframe
    ws_cf = wb.active
    ws_cf.title = "Codeframe"
    headers_cf = ["Kode", "Nett / Kategori", "Label (ID)", "Label (EN)"]
    hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top")

    for ci, h in enumerate(headers_cf, 1):
        cell = ws_cf.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

    current_nett = None
    nett_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    nett_font = Font(bold=True, name="Arial", size=10, color="1F4E78")
    body_font = Font(name="Arial", size=10)

    ri = 2
    for c in codeframe:
        if c['nett'] != current_nett:
            current_nett = c['nett']
            for ci in range(1, 5):
                cell = ws_cf.cell(row=ri, column=ci, value=current_nett if ci == 2 else "")
                cell.fill = nett_fill
                cell.font = nett_font
                cell.border = border
                cell.alignment = wrap
            ri += 1
        for ci, val in enumerate([c['code'], c['nett'], c['label_id'], c['label_en']], 1):
            cell = ws_cf.cell(row=ri, column=ci, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = center if ci == 1 else wrap
        ri += 1

    ws_cf.column_dimensions["A"].width = 8
    ws_cf.column_dimensions["B"].width = 25
    ws_cf.column_dimensions["C"].width = 40
    ws_cf.column_dimensions["D"].width = 40
    ws_cf.freeze_panes = "A2"

    # Sheet 2: Rawdata hasil
    ws_raw = wb.create_sheet("Rawdata")
    for ci, col in enumerate(df_result.columns, 1):
        cell = ws_raw.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

    code_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    for ri, row in enumerate(df_result.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws_raw.cell(row=ri, column=ci, value=val)
            col_name = df_result.columns[ci-1]
            cell.font = body_font
            cell.border = border
            cell.alignment = wrap
            if "Code" in col_name or "code" in col_name.lower():
                cell.fill = code_fill

    for ci in range(1, len(df_result.columns)+1):
        ws_raw.column_dimensions[get_column_letter(ci)].width = 30

    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(df_result.columns))}1"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────
for key in ["codeframe", "df_raw", "col_groups", "question_context",
            "autocode_results", "step", "api_key", "language"]:
    if key not in st.session_state:
        st.session_state[key] = None

if st.session_state.step is None:
    st.session_state.step = "upload"

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.image("https://img.icons8.com/fluency/96/artificial-intelligence.png", width=60)
    st.title("AI Verbatim Coder")
    st.caption("Powered by Gemini AI")
    st.divider()

    api_key = st.text_input(
        "🔑 Gemini API Key",
        type="password",
        value=st.session_state.api_key or "",
        placeholder="AIza..."
    )
    if api_key:
        st.session_state.api_key = api_key
        st.success("API Key tersimpan ✓")

    st.divider()

    mb = load_memory_bank()
    st.markdown(f"**📚 Memory Bank**")
    st.caption(f"{len(mb.get('studies', []))} studi tersimpan")
    for s in mb.get("studies", []):
        st.caption(f"• {s['study_id']}: {s.get('topic', '')[:40]}...")

    st.divider()

    with st.expander("➕ Tambah Studi ke Memory Bank"):
        st.caption("Upload file Excel hasil coding yang sudah selesai untuk memperkaya Memory Bank AI.")
        uploaded_memory = st.file_uploader("Upload Excel studi sebelumnya", type=["xlsx"], key="mem_upload")
        study_id_input = st.text_input("ID Studi (misal: Q5_2024)")
        topic_input = st.text_input("Topik singkat")
        question_input = st.text_area("Teks pertanyaan", height=60)
        verb_col_input = st.text_input("Nama kolom verbatim (contoh: Q5)")
        code_col_input = st.text_input("Nama kolom kode (contoh: Code)")

        if st.button("Simpan ke Memory Bank") and uploaded_memory and study_id_input:
            try:
                df_mem = pd.read_excel(uploaded_memory)
                samples = []
                if verb_col_input in df_mem.columns and code_col_input in df_mem.columns:
                    for _, row in df_mem.iterrows():
                        v = str(row[verb_col_input]).strip()
                        c = str(row[code_col_input]).strip()
                        if v and c and v != "nan" and c != "nan":
                            samples.append({"verbatim": v, "code": c})

                new_study = {
                    "study_id": study_id_input,
                    "topic": topic_input,
                    "question": question_input,
                    "type": "OE - dari upload user",
                    "codelist": [],
                    "verbatim_samples": samples[:50]
                }
                mb["studies"].append(new_study)
                save_memory_bank(mb)
                st.success(f"✓ Studi '{study_id_input}' disimpan! ({len(samples)} contoh)")
            except Exception as e:
                st.error(f"Error: {e}")

    if st.button("🔄 Reset / Mulai Ulang", use_container_width=True):
        for key in ["codeframe", "df_raw", "col_groups", "question_context",
                    "autocode_results", "step"]:
            st.session_state[key] = None
        st.session_state.step = "upload"
        st.rerun()

# ─────────────────────────────────────────────
# STEP INDICATOR
# ─────────────────────────────────────────────
steps = ["upload", "select_cols", "codeframe", "autocode", "done"]
step_labels = ["1️⃣ Upload", "2️⃣ Pilih Kolom", "3️⃣ Review Codeframe", "4️⃣ Autocode", "5️⃣ Download"]
current_step_idx = steps.index(st.session_state.step) if st.session_state.step in steps else 0

cols_step = st.columns(5)
for i, (col, label) in enumerate(zip(cols_step, step_labels)):
    with col:
        if i < current_step_idx:
            st.markdown(f"<div style='text-align:center;color:#27ae60;font-weight:bold'>{label} ✓</div>", unsafe_allow_html=True)
        elif i == current_step_idx:
            st.markdown(f"<div style='text-align:center;color:#2980b9;font-weight:bold;background:#EBF5FB;padding:4px;border-radius:6px'>{label}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div style='text-align:center;color:#bdc3c7'>{label}</div>", unsafe_allow_html=True)

st.divider()

# ─────────────────────────────────────────────
# STEP 1: UPLOAD
# ─────────────────────────────────────────────
if st.session_state.step == "upload":
    st.header("📂 Upload File Verbatim")

    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded = st.file_uploader("Upload file Excel (.xlsx) berisi data verbatim", type=["xlsx"])
        language = st.selectbox(
            "🌐 Bahasa output label",
            ["Bilingual (ID + EN)", "Indonesia saja", "Inggris saja"]
        )

    with col2:
        st.info("""
        **Format file yang didukung:**
        - Kolom verbatim bisa lebih dari 1
        - Bisa ada kolom Code yang masih kosong
        - Baris pertama = header
        - Sheet aktif = data
        """)

    if uploaded:
        try:
            df = pd.read_excel(uploaded)
            st.success(f"✓ File berhasil dibaca: {df.shape[0]} baris, {df.shape[1]} kolom")
            st.dataframe(df.head(5), use_container_width=True)

            if st.button("▶️ Lanjut Pilih Kolom", type="primary", use_container_width=True):
                st.session_state.df_raw = df
                st.session_state.language = language
                st.session_state.step = "select_cols"
                st.rerun()
        except Exception as e:
            st.error(f"Gagal membaca file: {e}")

# ─────────────────────────────────────────────
# STEP 2: PILIH KOLOM
# ─────────────────────────────────────────────
elif st.session_state.step == "select_cols":
    st.header("🗂️ Pilih Kolom yang Akan Di-coding")
    df = st.session_state.df_raw
    all_cols = list(df.columns)

    st.info("Setiap **Grup** = satu topik/kategori yang share 1 codeframe. Bisa tambah beberapa grup untuk topik berbeda dalam 1 file.")

    # Init builder
    if "col_groups_builder" not in st.session_state or st.session_state.col_groups_builder is None:
        st.session_state.col_groups_builder = []

    # Form tambah grup baru
    st.markdown("### ➕ Tambah Grup Baru")
    with st.container(border=True):
        grp_name = st.text_input("Nama Kategori/Nett", placeholder="Contoh: Alasan Memilih Brand", key="grp_name_input")
        grp_cols = st.multiselect("Pilih kolom verbatim untuk grup ini", all_cols, key="grp_cols_input")
        grp_ctx  = st.text_area(
            "Konteks pertanyaan untuk grup ini",
            placeholder="Contoh: Q5. Apa alasan Anda memilih brand ini? / Why did you choose this brand?",
            height=70,
            key="grp_ctx_input"
        )
        grp_mode = st.radio(
            "Mode generate codeframe",
            ["⚡ CEPAT — sampel 150 verbatim (cukup untuk data homogen)",
             "🔍 LENGKAP — baca semua verbatim per batch (lebih akurat, lebih lambat)"],
            key="grp_mode_input"
        )

        if st.button("✅ Tambah Grup", type="primary"):
            if not grp_name:
                st.error("Nama kategori wajib diisi!")
            elif not grp_cols:
                st.error("Pilih minimal 1 kolom verbatim!")
            elif not grp_ctx:
                st.error("Konteks pertanyaan wajib diisi agar AI lebih akurat!")
            else:
                st.session_state.col_groups_builder.append({
                    "name": grp_name,
                    "columns": grp_cols,
                    "context": grp_ctx,
                    "mode": "fast" if "CEPAT" in grp_mode else "full"
                })
                st.rerun()

    # Tampilkan grup yang sudah ditambahkan
    if st.session_state.col_groups_builder:
        st.markdown("### 📋 Grup yang Sudah Ditambahkan")
        for i, grp in enumerate(st.session_state.col_groups_builder):
            with st.container(border=True):
                col_x, col_y = st.columns([5, 1])
                with col_x:
                    mode_icon = "⚡" if grp.get("mode") == "fast" else "🔍"
                    st.markdown(f"**{grp['name']}** {mode_icon}")
                    st.caption(f"Kolom: {', '.join([f'`{c}`' for c in grp['columns']])}")
                    st.caption(f"Konteks: _{grp.get('context', '-')[:80]}..._")
                with col_y:
                    if st.button("🗑️", key=f"del_{i}"):
                        st.session_state.col_groups_builder.pop(i)
                        st.rerun()

        st.divider()
        if st.button("▶️ Generate Codeframe dengan AI", type="primary", use_container_width=True):
            if not st.session_state.api_key:
                st.error("Masukkan Gemini API Key di sidebar terlebih dahulu!")
            else:
                st.session_state.col_groups = st.session_state.col_groups_builder
                st.session_state.step = "codeframe"
                st.rerun()

# ─────────────────────────────────────────────
# STEP 3: GENERATE & REVIEW CODEFRAME
# ─────────────────────────────────────────────
elif st.session_state.step == "codeframe":
    st.header("🧠 Generate & Review Codeframe")
    df = st.session_state.df_raw
    mb = load_memory_bank()

    if st.session_state.codeframe is None:
        import random

        # Kumpulkan verbatim per grup
        all_verbatims = []
        for grp in st.session_state.col_groups:
            for col in grp["columns"]:
                if col in df.columns:
                    vals = df[col].dropna().astype(str).tolist()
                    all_verbatims.extend([v for v in vals if v.strip() and v != "nan"])

        # Gunakan context & mode dari grup pertama (atau gabungan semua grup)
        # Karena setiap grup punya codeframe sendiri, proses per grup
        grp_codeframes = {}
        all_success = True

        for g_idx, grp in enumerate(st.session_state.col_groups):
            grp_verbatims = []
            for col in grp["columns"]:
                if col in df.columns:
                    vals = df[col].dropna().astype(str).tolist()
                    grp_verbatims.extend([v for v in vals if v.strip() and v != "nan"])

            if not grp_verbatims:
                continue

            mode = grp.get("mode", "fast")
            ctx  = grp.get("context", grp["name"])

            if mode == "fast":
                # Sampel 150 acak
                sample = random.sample(grp_verbatims, min(150, len(grp_verbatims)))
                with st.spinner(f"⚡ [{grp['name']}] Mode CEPAT — membaca {len(sample)} sampel dari {len(grp_verbatims)} verbatim..."):
                    try:
                        time.sleep(2)
                        prompt = build_codeframe_prompt(mb, sample, ctx, st.session_state.language)
                        raw = call_gemini(st.session_state.api_key, prompt, temperature=0.3)
                        result = parse_json_response(raw)
                        grp_codeframes[grp["name"]] = result.get("codeframe", [])
                    except Exception as e:
                        st.error(f"Error grup '{grp['name']}': {e}")
                        all_success = False
            else:
                # Mode LENGKAP: baca semua per batch 100, kumpulkan tema, buat codeframe final
                batches = [grp_verbatims[i:i+100] for i in range(0, len(grp_verbatims), 100)]
                all_themes = []
                st.info(f"🔍 [{grp['name']}] Mode LENGKAP — memproses {len(batches)} batch dari {len(grp_verbatims)} verbatim...")
                progress_bar = st.progress(0)
                for b_i, batch in enumerate(batches):
                    try:
                        time.sleep(3)
                        prompt_batch = build_codeframe_prompt(mb, batch, ctx, st.session_state.language)
                        raw_batch = call_gemini(st.session_state.api_key, prompt_batch, temperature=0.3)
                        res_batch = parse_json_response(raw_batch)
                        all_themes.extend(res_batch.get("codeframe", []))
                        progress_bar.progress((b_i + 1) / len(batches))
                    except Exception as e:
                        st.warning(f"Batch {b_i+1} error (dilanjutkan): {e}")
                        time.sleep(5)

                # Gabung semua tema → minta AI buat codeframe final
                with st.spinner(f"🔍 [{grp['name']}] Menggabungkan {len(all_themes)} tema jadi codeframe final..."):
                    try:
                        themes_text = "\n".join([f"- {t.get('label_id','')} / {t.get('label_en','')}" for t in all_themes])
                        merge_prompt = f"""Kamu ahli coding verbatim riset pasar.
Berikut daftar tema mentah dari beberapa batch verbatim untuk pertanyaan: {ctx}

{themes_text}

Gabungkan tema yang serupa, hapus duplikat, dan buat CODEFRAME FINAL yang bersih dan terstruktur.
Ikuti konvensi penomoran (positif 100-an, negatif 500-an). Sertakan Nett/kategori besar.

Kembalikan HANYA JSON (tanpa teks lain):
{{"codeframe": [{{"code": 101, "nett": "Nama Nett", "label_id": "Label ID", "label_en": "Label EN"}}]}}"""
                        time.sleep(2)
                        raw_final = call_gemini(st.session_state.api_key, merge_prompt, temperature=0.2)
                        result_final = parse_json_response(raw_final)
                        grp_codeframes[grp["name"]] = result_final.get("codeframe", [])
                    except Exception as e:
                        st.error(f"Error merge codeframe '{grp['name']}': {e}")
                        grp_codeframes[grp["name"]] = all_themes
                        all_success = False

        if all_success or grp_codeframes:
            # Gabung semua codeframe dari semua grup (dengan marker nett per grup jika >1 grup)
            final_codeframe = []
            for grp_name, cf in grp_codeframes.items():
                if len(st.session_state.col_groups) > 1:
                    for item in cf:
                        item["nett"] = f"[{grp_name}] {item.get('nett','')}"
                final_codeframe.extend(cf)
            st.session_state.codeframe = final_codeframe
            st.session_state.grp_codeframes = grp_codeframes
            st.success(f"✅ Codeframe selesai! Total {len(final_codeframe)} kode dari {len(grp_codeframes)} grup. Silakan review di bawah.")

    st.warning("⚠️ **Harap review codeframe ini sebelum lanjut autocode!** Edit kode/label/nett jika ada yang perlu diperbaiki.")

    codeframe = st.session_state.codeframe

    # Editable codeframe table
    df_cf = pd.DataFrame(codeframe)
    if not df_cf.empty:
        st.markdown(f"**Total kode: {len(df_cf)}**")
        edited_df = st.data_editor(
            df_cf,
            use_container_width=True,
            num_rows="dynamic",
            column_config={
                "code": st.column_config.NumberColumn("Kode", width="small"),
                "nett": st.column_config.TextColumn("Nett / Kategori", width="medium"),
                "label_id": st.column_config.TextColumn("Label (ID)", width="large"),
                "label_en": st.column_config.TextColumn("Label (EN)", width="large"),
            }
        )

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("🔄 Generate Ulang Codeframe", use_container_width=True):
                st.session_state.codeframe = None
                st.rerun()
        with col_btn2:
            if st.button("✅ Setujui & Lanjut Autocode", type="primary", use_container_width=True):
                st.session_state.codeframe = edited_df.to_dict(orient="records")
                st.session_state.step = "autocode"
                st.rerun()

# ─────────────────────────────────────────────
# STEP 4: AUTOCODE
# ─────────────────────────────────────────────
elif st.session_state.step == "autocode":
    st.header("⚡ Autocode Verbatim")

    df = st.session_state.df_raw.copy()
    codeframe = st.session_state.codeframe
    col_groups = st.session_state.col_groups
    mb = load_memory_bank()
    api_key = st.session_state.api_key

    BATCH_SIZE = 25

    # Build result df: tambahkan kolom Code setelah setiap kolom verbatim
    result_cols = list(df.columns)
    insert_map = {}  # col_name -> hasil code list

    st.info(f"Akan melakukan autocode untuk {len(col_groups)} grup kolom. Proses ini mungkin memakan beberapa menit.")

    if st.session_state.autocode_results is None:
        if st.button("▶️ Mulai Autocode", type="primary", use_container_width=True):
            all_results = {}
            progress = st.progress(0)
            status = st.empty()

            total_cols = sum(len(g["columns"]) for g in col_groups)
            done_cols = 0

            for grp in col_groups:
                for col in grp["columns"]:
                    if col not in df.columns:
                        continue
                    verbatims = df[col].fillna("").astype(str).tolist()
                    codes_out = [""] * len(verbatims)

                    batches = [verbatims[i:i+BATCH_SIZE] for i in range(0, len(verbatims), BATCH_SIZE)]
                    for b_idx, batch in enumerate(batches):
                        status.markdown(f"🔄 Processing kolom **{col}** - batch {b_idx+1}/{len(batches)}...")
                        non_empty = [(i, v) for i, v in enumerate(batch) if v.strip() and v != "nan"]
                        if non_empty:
                            try:
                                batch_verbs = [v for _, v in non_empty]
                                prompt = build_autocode_prompt(mb, codeframe, batch_verbs, st.session_state.question_context)
                                raw = call_gemini(api_key, prompt, temperature=0.1)
                                results_batch = parse_json_response(raw)

                                global_offset = b_idx * BATCH_SIZE
                                for j, item in enumerate(results_batch):
                                    if j < len(non_empty):
                                        orig_idx, _ = non_empty[j]
                                        codes_out[global_offset + orig_idx] = str(item.get("code", ""))
                            except Exception as e:
                                status.error(f"Error batch {b_idx+1} kolom {col}: {e}")
                            time.sleep(1)  # rate limit safety

                    all_results[col] = codes_out
                    done_cols += 1
                    progress.progress(done_cols / total_cols)

            st.session_state.autocode_results = all_results
            status.success("✅ Autocode selesai!")
            st.rerun()

    else:
        st.success("✅ Autocode selesai! Preview hasil di bawah.")

        # Build result dataframe: insert Code column after each verbatim column
        df_out = st.session_state.df_raw.copy()
        insert_offset = 0
        for grp in col_groups:
            for col in grp["columns"]:
                if col in st.session_state.autocode_results:
                    col_idx = list(df_out.columns).index(col) + 1 + insert_offset
                    code_col_name = f"Code_{col}" if f"Code_{col}" not in df_out.columns else f"Code_{col}_AI"
                    df_out.insert(col_idx, code_col_name, st.session_state.autocode_results[col])
                    insert_offset += 1

        st.dataframe(df_out.head(20), use_container_width=True)

        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("🔄 Ulangi Autocode", use_container_width=True):
                st.session_state.autocode_results = None
                st.rerun()
        with col_b:
            if st.button("✅ Lanjut Download", type="primary", use_container_width=True):
                st.session_state.df_result = df_out
                st.session_state.step = "done"
                st.rerun()

# ─────────────────────────────────────────────
# STEP 5: DONE / DOWNLOAD
# ─────────────────────────────────────────────
elif st.session_state.step == "done":
    st.header("🎉 Selesai! Download Hasil")
    st.balloons()

    df_result = st.session_state.df_result
    codeframe = st.session_state.codeframe
    col_groups = st.session_state.col_groups

    st.success(f"✅ Berhasil! {len(df_result)} baris verbatim sudah di-code.")
    st.dataframe(df_result.head(10), use_container_width=True)

    excel_output = create_output_excel(df_result, codeframe, col_groups)

    st.download_button(
        label="⬇️ Download Excel Hasil (Codeframe + Rawdata)",
        data=excel_output,
        file_name="hasil_coding_verbatim.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )

    st.divider()

    with st.expander("💾 Simpan studi ini ke Memory Bank?"):
        st.caption("Dengan menyimpan hasil ini, AI akan semakin pintar untuk studi berikutnya.")
        new_study_id = st.text_input("ID Studi baru", placeholder="Contoh: D2_2025")
        new_topic = st.text_input("Topik singkat", placeholder="Contoh: Alasan memilih kendaraan Isuzu")

        if st.button("💾 Simpan ke Memory Bank") and new_study_id:
            mb = load_memory_bank()
            samples = []
            for grp in col_groups:
                for col in grp["columns"]:
                    code_col = f"Code_{col}"
                    if col in df_result.columns and code_col in df_result.columns:
                        for _, row in df_result.iterrows():
                            v = str(row[col]).strip()
                            c = str(row[code_col]).strip()
                            if v and c and v != "nan" and c != "nan":
                                samples.append({"verbatim": v, "code": c})

            new_study = {
                "study_id": new_study_id,
                "topic": new_topic,
                "question": st.session_state.question_context,
                "type": "OE - dari hasil autocode",
                "codelist": codeframe,
                "verbatim_samples": samples[:60]
            }
            mb["studies"].append(new_study)
            save_memory_bank(mb)
            st.success(f"✓ Studi '{new_study_id}' berhasil disimpan ke Memory Bank! ({len(samples)} contoh)")
